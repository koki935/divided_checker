"""スクリーニング結果をExcelに出力する"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path(__file__).parent.parent / "output"


def export_violations(results: list[dict], filename: str = "violations.xlsx") -> Path:
    """
    チェック結果をExcelに出力する。

    Args:
        results: check_violation の返り値に会社情報を付加した dict のリスト
        filename: 出力ファイル名

    Returns:
        出力ファイルのパス
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / filename

    wb = Workbook()

    # ===== シート1: 違反候補（HIGH/MEDIUM信頼度で超過あり） =====
    ws_violation = wb.active
    ws_violation.title = "違反候補"
    _write_sheet(
        ws_violation,
        [r for r in results if r.get("is_violation_candidate") and r.get("confidence") in ("HIGH", "MEDIUM")],
        highlight_color="FFCCCC",
    )

    # ===== シート2: 全件 =====
    ws_all = wb.create_sheet("全件")
    _write_sheet(ws_all, results)

    # ===== シート3: データ不足（要手動確認） =====
    ws_low = wb.create_sheet("データ不足")
    _write_sheet(
        ws_low,
        [r for r in results if r.get("confidence") == "LOW"],
        highlight_color="FFF2CC",
    )

    wb.save(output_path)
    return output_path


COLUMNS = [
    ("証券コード",       "sec_code",             10),
    ("会社名",           "name",                 20),
    ("EDINETコード",     "edinet_code",          12),
    ("決算期末",         "period_end",           12),
    ("提出年",           "fiscal_year",          8),
    ("分配可能額(円)",   "distributable_amount", 20),
    ("配当総額(円)",     "dividends_paid",       20),
    ("超過額(円)",       "excess",               20),
    ("超過率(%)",        "excess_rate",          10),
    ("違反候補",         "is_violation_candidate", 8),
    ("信頼度",           "confidence",           8),
    ("欠損項目",         "missing_fields",       30),
    ("配当タグ",         "dividends_paid_tag",   25),
]


def _write_sheet(ws, data: list[dict], highlight_color: str | None = None):
    """シートにデータを書き込む"""
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ヘッダー
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # データ行
    for row_idx, record in enumerate(data, 2):
        row_fill = None
        if highlight_color and record.get("is_violation_candidate"):
            row_fill = PatternFill("solid", fgColor=highlight_color)

        for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
            value = record.get(field)
            # 整形
            if isinstance(value, list):
                value = ", ".join(value)
            elif isinstance(value, bool):
                value = "YES" if value else "NO"
            elif isinstance(value, float) and field in ("distributable_amount", "dividends_paid", "excess"):
                value = int(value)  # 円は整数で

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            if row_fill:
                cell.fill = row_fill

            # 数値列は右寄せ・カンマ区切り
            if field in ("distributable_amount", "dividends_paid", "excess"):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif field == "excess_rate":
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal="right")

    # オートフィルター
    if data:
        ws.auto_filter.ref = ws.dimensions
